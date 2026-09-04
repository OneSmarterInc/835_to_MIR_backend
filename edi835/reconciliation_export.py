"""Professional Excel export for the filtered reconciliation result set."""

from collections import Counter
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY = "172033"
RED = "ED1C24"
TEAL = "087F8C"
LIGHT_BLUE = "EAF1F7"
WHITE = "FFFFFF"
GRAY = "52647B"
THIN_GRAY = Side(style="thin", color="D7E0EA")


def _money(value):
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal("0")


def build_reconciliation_workbook(*, client, rows, total, search="", status=""):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reconciliation Results"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A11"

    for column, width in enumerate((20, 18, 24, 20, 28, 28, 18, 18, 18, 22), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    sheet.merge_cells("A1:J1")
    sheet["A1"] = "MIR / RECON Reconciliation Results"
    sheet["A1"].font = Font(size=18, bold=True, color=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    generated = datetime.now(ZoneInfo("America/New_York")).strftime("%B %d, %Y at %I:%M:%S %p %Z")
    filter_text = search.strip() or "None"
    status_text = status.replace("_", " ").title() if status else "All statuses"
    statuses = Counter(row.get("status") or "UNKNOWN" for row in rows)
    matched = total - statuses.get("NOT_IN_MIR", 0) - statuses.get("NOT_IN_RECON", 0)
    total_mir = sum((_money(row.get("amount_to_pay")) for row in rows), Decimal("0"))
    total_recon = sum((_money(row.get("recon_paid_amount")) for row in rows), Decimal("0"))
    matched_difference = sum(
        (_money(row.get("difference_amount")) for row in rows if row.get("mir_claim_id") and row.get("recon_filename")),
        Decimal("0"),
    )

    client_name = "Global System"
    if client:
        client_name = f"{client.name} ({client.client_code})" if client.client_code else client.name
    summary = [
        ("Client", client_name, "Generated (EST)", generated),
        ("Search filter", filter_text, "Status filter", status_text),
        ("Matching claims", total, "Matched in MIR and RECON", matched),
        ("MIR-only claims", statuses.get("NOT_IN_RECON", 0), "RECON-only claims", statuses.get("NOT_IN_MIR", 0)),
        ("Total MIR amount", total_mir, "Total RECON amount", total_recon),
        ("Net difference (matched)", matched_difference, "Export scope", "All filtered entries"),
    ]
    for row_index, values in enumerate(summary, start=3):
        for offset, value in ((1, values[0]), (2, values[1]), (5, values[2]), (6, values[3])):
            cell = sheet.cell(row=row_index, column=offset, value=value)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center")
            if offset in (1, 5):
                cell.font = Font(size=10, bold=True, color=GRAY)
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            else:
                cell.font = Font(size=10, bold=isinstance(value, (int, Decimal)), color=NAVY)
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=4)
        sheet.merge_cells(start_row=row_index, start_column=6, end_row=row_index, end_column=9)
        for column in range(2, 5):
            sheet.cell(row_index, column).border = Border(bottom=THIN_GRAY)
        for column in range(6, 10):
            sheet.cell(row_index, column).border = Border(bottom=THIN_GRAY)

    currency_format = '$#,##0.00;[Red]-$#,##0.00'
    for cell in (sheet["B7"], sheet["F7"], sheet["B8"]):
        cell.number_format = currency_format

    headers = (
        "Highmark Claim Number", "Internal Claim Number", "Patient Name", "Member ID", "MIR File", "RECON File",
        "Amount in MIR", "Amount in RECON", "Difference", "Status",
    )
    header_row = 10
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 30

    for row_index, row in enumerate(rows, start=header_row + 1):
        matched_row = bool(row.get("mir_claim_id") and row.get("recon_filename"))
        values = (
            row.get("highmark_claim_number") or "-",
            row.get("internal_claim_number") or "-",
            row.get("patient_name") or "RECON-only claim",
            row.get("member_id") or "-",
            row.get("mir_filename") or "-",
            row.get("recon_filename") or "-",
            _money(row.get("amount_to_pay")),
            _money(row.get("recon_paid_amount")),
            _money(row.get("difference_amount")) if matched_row else "-",
            str(row.get("status") or "").replace("_", " ").title(),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.border = Border(bottom=Side(style="hair", color="D7E0EA"))
            cell.alignment = Alignment(vertical="top", wrap_text=column in (2, 3, 5, 6))
            if column in (7, 8) or (column == 9 and matched_row):
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal="right", vertical="top")
        if row_index % 2 == 0:
            for column in range(1, 11):
                sheet.cell(row_index, column).fill = PatternFill("solid", fgColor="F7F9FC")

    last_row = max(header_row + 1, header_row + len(rows))
    if not rows:
        sheet.cell(header_row + 1, 1, "No reconciliation results match the selected filters.")
        sheet.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=10)
        sheet.cell(header_row + 1, 1).alignment = Alignment(horizontal="center")
    else:
        table = Table(displayName="ReconciliationResults", ref=f"A{header_row}:J{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)

    # The table already owns the AutoFilter for populated exports. Adding a\n    # worksheet AutoFilter over the same range creates overlapping OOXML that\n    # desktop Excel reports as damaged and attempts to repair on open.\n    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = f"1:{header_row}"
    sheet.oddFooter.center.text = "OneSmarter Confidential · Page &P of &N"
    sheet.oddFooter.center.size = 9
    sheet.oddFooter.center.color = GRAY

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
